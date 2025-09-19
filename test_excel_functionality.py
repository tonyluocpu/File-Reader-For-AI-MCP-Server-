#!/usr/bin/env python3
"""
Comprehensive test script for Excel functionality in AI File Bridge
Tests formulas, formatting, multiple sheets, and data validation
"""

import json
import sys
from pathlib import Path
from excel_reader import ExcelReader
from word_reader import WordReader

def create_test_excel_file():
    """Create a test Excel file with various features"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.formatting.rule import DataBarRule
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Data"
        
        # Add headers with formatting
        headers = ["Name", "Age", "Department", "Salary", "Bonus", "Total", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data
        data = [
            ["Alice Johnson", 28, "Engineering", 75000, 5000, "=D2+E2", "Active"],
            ["Bob Smith", 35, "Marketing", 65000, 3000, "=D3+E3", "Active"],
            ["Carol Davis", 42, "Sales", 80000, 8000, "=D4+E4", "Active"],
            ["David Wilson", 29, "Engineering", 72000, 4000, "=D5+E5", "Inactive"],
            ["Eva Brown", 31, "HR", 60000, 2000, "=D6+E6", "Active"]
        ]
        
        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # Add conditional formatting for salary column
                if col == 4:  # Salary column
                    if row == 2 or row == 4:  # High salaries
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        
        # Add data validation for Status column
        dv = DataValidation(type="list", formula1='"Active,Inactive"')
        dv.add("G2:G6")
        ws.add_data_validation(dv)
        
        # Add a second worksheet with charts data
        ws2 = wb.create_sheet("Charts Data")
        chart_data = [
            ["Month", "Revenue", "Expenses", "Profit"],
            ["Jan", 100000, 75000, "=B2-C2"],
            ["Feb", 120000, 80000, "=B3-C3"],
            ["Mar", 110000, 78000, "=B4-C4"],
            ["Apr", 130000, 85000, "=B5-C5"],
            ["May", 125000, 82000, "=B6-C6"]
        ]
        
        for row, row_data in enumerate(chart_data, 1):
            for col, value in enumerate(row_data, 1):
                ws2.cell(row=row, column=col, value=value)
        
        # Format the second sheet
        for col in range(1, 5):
            cell = ws2.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Save the test file
        test_file = "test_excel_features.xlsx"
        wb.save(test_file)
        print(f"✅ Created test Excel file: {test_file}")
        return test_file
        
    except ImportError:
        print("❌ openpyxl not available for creating test file")
        return None

def test_excel_reader():
    """Test Excel reader functionality"""
    print("🧪 Testing Excel Reader Functionality")
    print("=" * 50)
    
    # Create test file
    test_file = create_test_excel_file()
    if not test_file:
        print("❌ Cannot test without test file")
        return
    
    reader = ExcelReader()
    
    try:
        # Test basic reading
        print("\n📊 Testing Basic Excel Reading...")
        result = reader.read_excel_file(test_file, include_formatting=True, include_formulas=True, include_validation=True)
        
        if result.get("success"):
            print("✅ Excel file read successfully!")
            print(f"📁 File: {result['file_path']}")
            print(f"📋 File Type: {result['file_type']}")
            print(f"📊 Total Sheets: {result['workbook_info']['total_sheets']}")
            print(f"📝 Sheet Names: {result['workbook_info']['sheet_names']}")
            
            # Test each worksheet
            for sheet in result['worksheets']:
                print(f"\n📄 Sheet: {sheet['name']}")
                print(f"   📏 Dimensions: {sheet['dimensions']['used_range']}")
                print(f"   📊 Total Cells: {sheet['total_cells']}")
                print(f"   🧮 Formulas Found: {len(sheet['formulas'])}")
                print(f"   🎨 Formatted Cells: {len(sheet['formatting'])}")
                print(f"   ✅ Data Validation Rules: {len(sheet['data_validation'])}")
                
                # Show sample formulas
                if sheet['formulas']:
                    print("   🧮 Sample Formulas:")
                    for i, (cell_ref, formula_data) in enumerate(sheet['formulas'].items()):
                        if i < 3:  # Show first 3 formulas
                            print(f"      {cell_ref}: {formula_data['formula']} = {formula_data['calculated_value']}")
                
                # Show sample formatting
                if sheet['formatting']:
                    print("   🎨 Sample Formatting:")
                    for i, (cell_ref, formatting) in enumerate(sheet['formatting'].items()):
                        if i < 2:  # Show first 2 formatted cells
                            if formatting.get('font', {}).get('bold'):
                                print(f"      {cell_ref}: Bold text")
                            if formatting.get('fill'):
                                print(f"      {cell_ref}: Background color")
                
                # Show sample data validation
                if sheet['data_validation']:
                    print("   ✅ Data Validation:")
                    for cell_range, validation in sheet['data_validation'].items():
                        print(f"      {cell_range}: {validation['type']} validation")
            
            print("\n🎯 Excel Reading Test: PASSED ✅")
            
        else:
            print(f"❌ Excel reading failed: {result.get('error')}")
            
    except Exception as e:
        print(f"❌ Test failed with error: {str(e)}")
    
    finally:
        # Clean up test file
        if Path(test_file).exists():
            Path(test_file).unlink()
            print(f"🧹 Cleaned up test file: {test_file}")

def test_word_reader():
    """Test Word reader functionality"""
    print("\n📝 Testing Word Reader Functionality")
    print("=" * 50)
    
    reader = WordReader()
    
    # Test with existing Word file if available
    test_files = ["../fiction/学园轻曲 (1).docx", "sample.docx", "test.docx"]
    test_file = None
    
    for file_path in test_files:
        if Path(file_path).exists():
            test_file = file_path
            break
    
    if not test_file:
        print("❌ No Word document found for testing")
        return
    
    try:
        print(f"📄 Testing with: {test_file}")
        result = reader.read_word_document(test_file, include_formatting=True)
        
        if result.get("success"):
            print("✅ Word document read successfully!")
            print(f"📁 File: {result['file_path']}")
            print(f"📋 File Type: {result['file_type']}")
            print(f"📝 Total Paragraphs: {result['total_paragraphs']}")
            print(f"📊 Total Tables: {result['total_tables']}")
            print(f"🎨 Formatting Included: {result['formatting_included']}")
            
            # Show sample paragraphs
            if result['paragraphs']:
                print("\n📝 Sample Paragraphs:")
                for i, para in enumerate(result['paragraphs'][:3]):  # First 3 paragraphs
                    text = para['text'][:100] + "..." if len(para['text']) > 100 else para['text']
                    print(f"   {i+1}. {text}")
            
            print("\n🎯 Word Reading Test: PASSED ✅")
            
        else:
            print(f"❌ Word reading failed: {result.get('error')}")
            
    except Exception as e:
        print(f"❌ Test failed with error: {str(e)}")

def test_mcp_server():
    """Test the MCP server functionality"""
    print("\n🔧 Testing MCP Server Functionality")
    print("=" * 50)
    
    try:
        from ai_file_bridge_server import AIFileBridgeServer
        
        server = AIFileBridgeServer()
        
        # Test supported formats
        print("📋 Supported Formats:")
        for ext, desc in server.supported_formats.items():
            print(f"   {ext}: {desc}")
        
        print(f"\n🎯 MCP Server Test: PASSED ✅")
        
    except Exception as e:
        print(f"❌ MCP Server test failed: {str(e)}")

def main():
    """Run all tests"""
    print("🚀 AI File Bridge - Comprehensive Functionality Test")
    print("=" * 60)
    
    # Test Excel functionality
    test_excel_reader()
    
    # Test Word functionality  
    test_word_reader()
    
    # Test MCP server
    test_mcp_server()
    
    print("\n🎉 All Tests Completed!")
    print("=" * 60)

if __name__ == "__main__":
    main()
