#!/usr/bin/env python3
"""
Excel file reader module for AI File Bridge
Supports .xlsx, .xls files with comprehensive data extraction
"""

import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from datetime import datetime
import json

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Border, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import xlrd
except ImportError:
    print("xlrd not installed. Install with: pip install xlrd", file=sys.stderr)
    sys.exit(1)


class ExcelReader:
    """Comprehensive Excel file reader with formula and formatting extraction"""
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls', '.xlsm']
    
    def read_excel_file(self, file_path: str, include_formatting: bool = True, 
                       include_formulas: bool = True, include_validation: bool = True) -> Dict[str, Any]:
        """
        Read Excel file with comprehensive data extraction
        
        Args:
            file_path: Path to Excel file
            include_formatting: Whether to extract cell formatting
            include_formulas: Whether to extract formulas
            include_validation: Whether to extract data validation rules
            
        Returns:
            Dictionary with extracted Excel data
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if path.suffix.lower() not in self.supported_formats:
            raise ValueError(f"Unsupported Excel format: {path.suffix}. Supported: {self.supported_formats}")
        
        try:
            if path.suffix.lower() in ['.xlsx', '.xlsm']:
                return self._read_xlsx_file(path, include_formatting, include_formulas, include_validation)
            elif path.suffix.lower() == '.xls':
                return self._read_xls_file(path, include_formatting, include_formulas, include_validation)
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "file_path": str(path)
            }
    
    def _read_xlsx_file(self, file_path: Path, include_formatting: bool, 
                       include_formulas: bool, include_validation: bool) -> Dict[str, Any]:
        """Read .xlsx/.xlsm files using openpyxl"""
        
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        result = {
            "success": True,
            "file_path": str(file_path),
            "file_type": "Excel Workbook (.xlsx/.xlsm)",
            "worksheets": [],
            "workbook_info": {
                "total_sheets": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "creator": getattr(workbook.properties, 'creator', 'Unknown'),
                "created": workbook.properties.created.isoformat() if workbook.properties.created else None,
                "modified": workbook.properties.modified.isoformat() if workbook.properties.modified else None
            }
        }
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_data = self._extract_worksheet_data(
                worksheet, include_formatting, include_formulas, include_validation
            )
            result["worksheets"].append(sheet_data)
        
        return result
    
    def _read_xls_file(self, file_path: Path, include_formatting: bool, 
                      include_formulas: bool, include_validation: bool) -> Dict[str, Any]:
        """Read .xls files using xlrd"""
        
        workbook = xlrd.open_workbook(str(file_path), formatting_info=include_formatting)
        
        result = {
            "success": True,
            "file_path": str(file_path),
            "file_type": "Excel Workbook (.xls)",
            "worksheets": [],
            "workbook_info": {
                "total_sheets": workbook.nsheets,
                "sheet_names": workbook.sheet_names(),
                "created": None,  # xlrd doesn't provide this info easily
                "modified": None
            }
        }
        
        for i in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(i)
            sheet_data = self._extract_xls_worksheet_data(
                worksheet, include_formatting, include_formulas, include_validation
            )
            result["worksheets"].append(sheet_data)
        
        return result
    
    def _extract_worksheet_data(self, worksheet, include_formatting: bool, 
                               include_formulas: bool, include_validation: bool) -> Dict[str, Any]:
        """Extract data from openpyxl worksheet"""
        
        # Get worksheet dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Extract cells data
        cells_data = []
        formulas = {}
        formatting_info = {}
        validation_rules = {}
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_ref = f"{get_column_letter(col)}{row}"
                
                # Basic cell data
                cell_data = {
                    "reference": cell_ref,
                    "row": row,
                    "column": col,
                    "value": self._get_cell_value(cell),
                    "data_type": str(type(cell.value).__name__)
                }
                
                # Extract formulas if requested
                if include_formulas and cell.data_type == 'f':
                    formulas[cell_ref] = {
                        "formula": cell.value,
                        "calculated_value": self._get_cell_value(cell)
                    }
                
                # Extract formatting if requested
                if include_formatting:
                    formatting_info[cell_ref] = self._extract_cell_formatting(cell)
                
                # Only include non-empty cells
                if cell.value is not None and str(cell.value).strip():
                    cells_data.append(cell_data)
        
        # Extract data validation rules
        if include_validation:
            validation_rules = self._extract_validation_rules(worksheet)
        
        return {
            "name": worksheet.title,
            "dimensions": {
                "max_row": max_row,
                "max_column": max_col,
                "used_range": f"A1:{get_column_letter(max_col)}{max_row}"
            },
            "cells": cells_data,
            "total_cells": len(cells_data),
            "formulas": formulas if include_formulas else {},
            "formatting": formatting_info if include_formatting else {},
            "data_validation": validation_rules if include_validation else {}
        }
    
    def _extract_xls_worksheet_data(self, worksheet, include_formatting: bool, 
                                   include_formulas: bool, include_validation: bool) -> Dict[str, Any]:
        """Extract data from xlrd worksheet"""
        
        cells_data = []
        formulas = {}
        formatting_info = {}
        
        for row in range(worksheet.nrows):
            for col in range(worksheet.ncols):
                cell_value = worksheet.cell_value(row, col)
                
                # Skip empty cells
                if cell_value == '' or cell_value is None:
                    continue
                
                cell_ref = f"{get_column_letter(col + 1)}{row + 1}"
                
                cell_data = {
                    "reference": cell_ref,
                    "row": row + 1,
                    "column": col + 1,
                    "value": cell_value,
                    "data_type": worksheet.cell_type(row, col)
                }
                
                # Extract formulas (xlrd limitation: formulas are not easily accessible)
                if include_formulas and worksheet.cell_type(row, col) == xlrd.XL_CELL_FORMULA:
                    formulas[cell_ref] = {
                        "formula": "Formula extraction not fully supported in .xls files",
                        "calculated_value": cell_value
                    }
                
                # Extract formatting
                if include_formatting:
                    formatting_info[cell_ref] = self._extract_xls_cell_formatting(worksheet, row, col)
                
                cells_data.append(cell_data)
        
        return {
            "name": worksheet.name,
            "dimensions": {
                "max_row": worksheet.nrows,
                "max_column": worksheet.ncols,
                "used_range": f"A1:{get_column_letter(worksheet.ncols)}{worksheet.nrows}"
            },
            "cells": cells_data,
            "total_cells": len(cells_data),
            "formulas": formulas if include_formulas else {},
            "formatting": formatting_info if include_formatting else {},
            "data_validation": {}  # xlrd doesn't support data validation extraction
        }
    
    def _get_cell_value(self, cell) -> Any:
        """Get cell value, handling different data types"""
        if cell.value is None:
            return None
        
        # Handle datetime objects
        if isinstance(cell.value, datetime):
            return cell.value.isoformat()
        
        return cell.value
    
    def _extract_cell_formatting(self, cell) -> Dict[str, Any]:
        """Extract cell formatting information"""
        formatting = {
            "font": {},
            "fill": {},
            "border": {},
            "alignment": {},
            "number_format": cell.number_format
        }
        
        # Font formatting
        if cell.font:
            formatting["font"] = {
                "name": cell.font.name,
                "size": cell.font.size,
                "bold": cell.font.bold,
                "italic": cell.font.italic,
                "underline": cell.font.underline,
                "color": str(cell.font.color.rgb) if cell.font.color and cell.font.color.rgb else None
            }
        
        # Fill (background) formatting
        if cell.fill:
            if isinstance(cell.fill.fill_type, PatternFill):
                formatting["fill"] = {
                    "pattern_type": cell.fill.fill_type.patternType,
                    "fg_color": str(cell.fill.fill_type.fgColor.rgb) if cell.fill.fill_type.fgColor and cell.fill.fill_type.fgColor.rgb else None,
                    "bg_color": str(cell.fill.fill_type.bgColor.rgb) if cell.fill.fill_type.bgColor and cell.fill.fill_type.bgColor.rgb else None
                }
        
        # Border formatting
        if cell.border:
            border_info = {}
            for side in ['top', 'bottom', 'left', 'right']:
                border_side = getattr(cell.border, side)
                if border_side:
                    border_info[side] = {
                        "style": border_side.style,
                        "color": str(border_side.color.rgb) if border_side.color and border_side.color.rgb else None
                    }
            formatting["border"] = border_info
        
        # Alignment formatting
        if cell.alignment:
            formatting["alignment"] = {
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "wrap_text": cell.alignment.wrap_text,
                "text_rotation": cell.alignment.text_rotation
            }
        
        return formatting
    
    def _extract_xls_cell_formatting(self, worksheet, row: int, col: int) -> Dict[str, Any]:
        """Extract formatting from xlrd worksheet (limited support)"""
        # xlrd has limited formatting support
        xf_index = worksheet.cell_xf_index(row, col)
        
        return {
            "xf_index": xf_index,
            "note": "Limited formatting support for .xls files"
        }
    
    def _extract_validation_rules(self, worksheet) -> Dict[str, Any]:
        """Extract data validation rules from worksheet"""
        validation_rules = {}
        
        # Note: openpyxl has limited data validation extraction capabilities
        # This is a simplified implementation
        for validation in worksheet.data_validations.dataValidation:
            validation_rules[str(validation.ranges)] = {
                "type": validation.type,
                "operator": validation.operator,
                "formula1": validation.formula1,
                "formula2": validation.formula2,
                "show_error_message": validation.showErrorMessage,
                "error_message": validation.error,
                "show_input_message": validation.showInputMessage,
                "input_message": validation.prompt
            }
        
        return validation_rules


def main():
    """Test function for Excel reader"""
    reader = ExcelReader()
    
    # Test with a sample file if available
    test_file = "sample.xlsx"
    if Path(test_file).exists():
        result = reader.read_excel_file(test_file)
        print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
    else:
        print(f"Test file {test_file} not found. Create a sample Excel file to test.")


if __name__ == "__main__":
    main()
